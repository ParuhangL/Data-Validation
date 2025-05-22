import pandas as pd
import tkinter as tk
from tkinter import messagebox
from tkinter.filedialog import askopenfilename
from nepali_datetime import date as NepaliDate
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import sys

# Initialize GUI window
root = tk.Tk()
root.withdraw()

# File paths
input_path = askopenfilename(
    title="Select the Excel file you want to validate",
    filetypes=[("Excel files", "*.xlsx *.xls")]
)
if not input_path:
    messagebox.showinfo("No file selected", "No file was selected. Exiting the program.")
    exit()

output_path = os.path.splitext(input_path)[0] + "_validated.xlsx"

# If output already exists, try to remove it
if os.path.exists(output_path):
    try:
        os.remove(output_path)
    except PermissionError:
        messagebox.showerror("File Error", f"Please close the file:\n{output_path}\nThen try again.")
        sys.exit()

# Load Excel
df = pd.read_excel(input_path)

# Clean column names and string values
df.columns = df.columns.str.strip()
for col in df.select_dtypes(include='object'):
    df[col] = df[col].map(lambda x: x.strip() if isinstance(x, str) else x)

# Member ID Validation
id_col = 'Membe Id'
df[id_col] = df[id_col].fillna(0).astype(int)
expected_ids = list(range(1, len(df) + 1))
actual_ids = df[id_col].tolist()

if actual_ids != expected_ids:
    msg = "The Membe ID column is not serial! \n\n"
    for i, (expected, actual) in enumerate(zip(expected_ids, actual_ids), start=1):
        df[id_col] = expected_ids
    messagebox.showwarning("Validation Alert", msg + "\n IDs have been re-sequenced and saved to final output.")
else:
    messagebox.showinfo("Validation Complete", "The 'Membe Id' column is already in proper serial order.")

# Automatically remove rows where both date and corresponding balance/amount are empty
date_balance_pairs = []

for col in df.columns:
    if 'date' in col.lower():
        for keyword in ['balance', 'amount']:
            possible_balance = next((b for b in df.columns if keyword in b.lower()), None)
            if possible_balance:
                date_balance_pairs.append((col, possible_balance))
                break

#drop rows where both are empty
for date_col, balance_col in date_balance_pairs:
    initial_len = len(df)
    df = df[~((df[date_col].isna() | (df[date_col].astype(str).str.strip() == "")) &
              (df[balance_col].isna() | (df[balance_col].astype(str).str.strip() == "")))]
    removed_rows = initial_len - len(df)
    if removed_rows > 0:
        print(f"Removed {removed_rows} rows where both '{date_col}' and '{balance_col}' were empty.")


# Detect potential date columns
potential_date_columns = [
    col for col in df.columns
    if 'date' in col.lower() and not col.lower().endswith('errors')
]

# Detect potential balance columns
potential_balance_columns = [
    col for col in df.columns
    if any(k in col.lower() for k in ['balance', 'amount', 'closing'])
]

# Date format and validation
current_bs_year = NepaliDate.today().year
ambiguous_format = re.compile(r'^\d{1,2}/\d{1,2}/\d{4}$')

date_errors = {}
balance_errors = {}
general_errors = {}

for date_col in potential_date_columns:
    for balance_col in potential_balance_columns:
        fixed_dates = []
        date_errs = []
        balance_errs = []
        general_errs = []

        for idx, row in df.iterrows():
            val = str(row[date_col]).strip() if pd.notna(row[date_col]) else ""
            balance_val = row[balance_col]
            date_error = ""
            balance_error = ""
            general_error = ""

            has_date = bool(val and val.lower() != 'nan')
            has_balance = pd.notna(balance_val) and str(balance_val).strip() not in ['', 'nan', 'NaN']

            if not has_date and not has_balance:
                general_error = f"Date and Balance missing for '{date_col}' and '{balance_col}'"
            elif has_date and not has_balance:
                balance_error = f"Balance missing for '{balance_col}'"
            elif has_balance and not has_date:
                date_error = f"Date missing for '{date_col}'"
            elif has_date and has_balance:
                val = val.replace('.', '/').replace('-', '/')
                if ambiguous_format.match(val):
                    date_error = f"Ambiguous format: '{val}'"
                else:
                    parts = val.split('/')
                    if len(parts) != 3:
                        date_error = f"Invalid format: '{val}'"
                    else:
                        try:
                            y, m, d = map(int, parts)
                            if y < 1900 or y > current_bs_year:
                                date_error += f"Invalid year: {y}. "
                            if m < 1 or m > 12:
                                date_error += f"Invalid month: {m}. "
                            if d < 1 or d > 32:
                                date_error += f"Invalid day: {d}. "
                            if not date_error:
                                val = f"{y:04d}/{m:02d}/{d:02d}"
                        except:
                            date_error = f"Non-numeric date: '{val}'"

            fixed_dates.append(val if val else "")
            date_errs.append(date_error)
            balance_errs.append(balance_error)
            general_errs.append(general_error if 'general_error' in locals() else "")

        df[date_col] = fixed_dates
        df[f"{date_col} Errors"] = date_errs
        df[f"{balance_col} Errors"] = balance_errs
        df[f"General Errors ({date_col} & {balance_col})"] = general_errs

        date_errors[date_col] = [(i+2, e) for i, e in enumerate(date_errs) if e]
        balance_errors[balance_col] = [(i+2, e) for i, e in enumerate(balance_errs) if e]
        general_errors[(date_col, balance_col)] = [(i + 2, e) for i, e in enumerate(general_errs) if e]


# Save validated file
df.to_excel(output_path, index=False)

# Load workbook and highlight errors
wb = load_workbook(output_path)
ws = wb.active

red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")

# Highlight date errors
for col in date_errors:
    base_col_idx = list(df.columns).index(col)
    err_col_idx = list(df.columns).index(f"{col} Errors")
    for row_idx, msg in date_errors[col]:
        ws.cell(row=row_idx, column=base_col_idx + 1).fill = red_fill
        ws.cell(row=row_idx, column=err_col_idx + 1).fill = yellow_fill

# Highlight balance errors
for col in balance_errors:
    base_col_idx = list(df.columns).index(col)
    err_col_idx = list(df.columns).index(f"{col} Errors")
    for row_idx, msg in balance_errors[col]:
        ws.cell(row=row_idx, column=base_col_idx + 1).fill = red_fill
        ws.cell(row=row_idx, column=err_col_idx + 1).fill = yellow_fill

#Hightlight general errors
for (date_col, balance_col), errs in general_errors.items():
    general_col_name = f"General Errors ({date_col} & {balance_col})"
    if general_col_name in df.columns:
        gen_err_idx = list(df.columns).index(general_col_name)
        for row_idx, msg in errs:
            # Highlight the general error cell yellow
            ws.cell(row=row_idx, column=gen_err_idx + 1).fill = yellow_fill
            # Highlight entire row light red
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = red_fill

# Save final workbook
wb.save(output_path)

# Display summary of errors
all_date_errors = sum(len(v) for v in date_errors.values())
all_balance_errors = sum(len(v) for v in balance_errors.values())
all_general_errors = sum(len(v) for v in general_errors.values())
total_errors = all_date_errors + all_balance_errors + all_general_errors

if total_errors:
    full_msg = "Validation found issues:\n\n"
    for col, errs in date_errors.items():
        full_msg += f"{col} (Date Errors): {len(errs)}\n"
    for col, errs in balance_errors.items():
        full_msg += f"{col} (Balance Errors): {len(errs)}\n"
    for (date_col, balance_col), errs in general_errors.items():
        full_msg += f"{date_col} & {balance_col} (General Errors): {len(errs)}\n"
    messagebox.showerror("Validation Issues", full_msg)
else:
    messagebox.showinfo("Validation Complete", "All date and balance values are valid.")
